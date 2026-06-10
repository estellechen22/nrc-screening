"""Excel 报告导出"""
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
EXPORT_DIR = os.path.join(DATA_DIR, "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

# 样式定义
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
GOOD_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
DEFAULT_FONT = Font(name="Microsoft YaHei", size=10)
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def export_single_to_excel(screen_result: dict, creator: dict, layers: list[dict]) -> str:
    """导出单个创作者筛查报告"""
    wb = Workbook()
    ws = wb.active
    ws.title = "筛查报告"

    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"] = f"创作者背调筛查报告 — {creator.get('name', '') or creator.get('handle', '')}"
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="1A1A2E")
    ws["A1"].alignment = Alignment(horizontal="center")

    # 基本信息
    row = 3
    info_items = [
        ("平台", creator.get("platform", "")),
        ("URL", creator.get("url", "")),
        ("粉丝数", str(creator.get("subs", ""))),
        ("国家/地区", creator.get("country", "")),
    ]
    for label, value in info_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # 综合评分
    row += 1
    ws.cell(row=row, column=1, value="综合评分").font = Font(bold=True, size=14)
    ws.cell(row=row, column=2, value=screen_result.get("composite_score", "N/A")).font = Font(size=14, bold=True)
    row += 1
    ws.cell(row=row, column=1, value="判定").font = Font(bold=True)
    verdict_map = {"approve": "✅ 通过", "review": "🟡 需审查", "reject": "🔴 不推荐"}
    ws.cell(row=row, column=2, value=verdict_map.get(screen_result.get("verdict", ""), "N/A"))

    # 一票否决
    veto_flags = screen_result.get("veto_flags", [])
    if isinstance(veto_flags, str):
        try:
            veto_flags = json.loads(veto_flags)
        except (json.JSONDecodeError, TypeError):
            veto_flags = []
    if veto_flags:
        row += 2
        ws.cell(row=row, column=1, value="一票否决").font = Font(bold=True, color="FF0000")
        for flag in veto_flags:
            row += 1
            ws.cell(row=row, column=1, value=flag.get("text", "") if isinstance(flag, dict) else str(flag))

    # 六层详情
    row += 2
    ws.cell(row=row, column=1, value="逐层筛查详情").font = Font(bold=True, size=12)
    row += 1

    headers = ["层级", "评分", "风险等级", "关键发现"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for layer in layers:
        details = layer.get("details", [])
        if isinstance(details, str):
            try:
                details = json.loads(details)
            except (json.JSONDecodeError, TypeError):
                details = []

        findings = []
        for section in details:
            for item in section.get("items", []):
                label = item.get("label", "")
                value = item.get("value", "")
                findings.append(f"{label}: {value}")

        ws.cell(row=row, column=1, value=f"L{layer['layer_number']}: {layer['layer_name']}").font = DEFAULT_FONT
        ws.cell(row=row, column=2, value=layer.get("score", "N/A")).font = DEFAULT_FONT
        ws.cell(row=row, column=3, value=layer.get("level", "")).font = DEFAULT_FONT
        ws.cell(row=row, column=4, value="\n".join(findings[:8])).font = DEFAULT_FONT
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)

        # 行颜色
        fill = GOOD_FILL if layer.get("score", 0) and layer["score"] >= 70 else (
            WARN_FILL if layer.get("score", 0) and layer["score"] >= 50 else BAD_FILL
        )
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = thin_border

        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 60

    filepath = os.path.join(EXPORT_DIR, f"report_{screen_result['id']}.xlsx")
    wb.save(filepath)
    return filepath


def export_batch_to_excel(batch_id: str, items: list[dict]) -> str:
    """导出批量筛查汇总表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "批量筛查汇总"

    # 标题
    ws.merge_cells("A1:H1")
    ws["A1"] = f"创作者批量背调筛查报告 — 批次 {batch_id}"
    ws["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="1A1A2E")

    # 表头
    headers = ["创作者", "平台", "粉丝数", "国家", "综合评分", "判定", "一票否决", "URL"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 数据行
    row = 4
    for item in items:
        veto = item.get("veto_flags", [])
        if isinstance(veto, str):
            try:
                veto = json.loads(veto)
            except (json.JSONDecodeError, TypeError):
                veto = []
        veto_text = "; ".join(
            [f.get("text", "") for f in veto if isinstance(f, dict)]
        ) if veto else ""

        verdict_map = {"approve": "✅ 通过", "review": "🟡 审查", "reject": "🔴 拒绝"}
        verdict = verdict_map.get(item.get("verdict", ""), item.get("status", ""))

        score = item.get("composite_score")
        fill = GOOD_FILL if (score and score >= 80) else (WARN_FILL if (score and score >= 60) else BAD_FILL)

        data = [
            item.get("name") or item.get("handle", "Unknown"),
            item.get("platform", ""),
            item.get("subs", ""),
            item.get("country", ""),
            score,
            verdict,
            veto_text,
            item.get("url", ""),
        ]
        for i, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.font = DEFAULT_FONT
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=(i == 7))
        row += 1

    # 列宽
    widths = [20, 10, 12, 10, 10, 12, 40, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filepath = os.path.join(EXPORT_DIR, f"batch_{batch_id}.xlsx")
    wb.save(filepath)
    return filepath
